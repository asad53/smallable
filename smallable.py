from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.touch_actions import TouchActions
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from fake_useragent import UserAgent
import time
import openpyxl
import math
from bs4 import BeautifulSoup as soup


def configure_driver():
    # Add additional Options to the webdriver
    chrome_options = Options()
    ua = UserAgent()
    userAgent = ua.random  # THIS IS FAKE AGENT IT WILL GIVE YOU NEW AGENT EVERYTIME
    print(userAgent)
    # add the argument and make the browser Headless.
    # chrome_options.add_argument("--headless")                    if you don't want to see the display on chrome just uncomment this
    chrome_options.add_argument(f'user-agent={userAgent}')  # useragent added
    chrome_options.add_argument("--log-level=3")  # removes error/warning/info messages displayed on the console
    chrome_options.add_argument("--disable-notifications")  # disable notifications
    chrome_options.add_argument(
        "--disable-infobars")  # disable infobars ""Chrome is being controlled by automated test software"  Although is isn't supported by Chrome anymore
    chrome_options.add_argument("start-maximized")  # will maximize chrome screen
    chrome_options.add_argument('--disable-gpu')  # disable gpu (not load pictures fully)
    chrome_options.add_argument("--disable-extensions")  # will disable developer mode extensions
    # chrome_options.add_argument('--proxy-server=%s' % PROXY)
    # chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
    prefs = {"profile.managed_default_content_settings.images": 2}
    chrome_options.add_experimental_option("prefs", prefs)             #we have disabled pictures (so no time is wasted in loading them)
    driver = webdriver.Chrome(ChromeDriverManager().install(),options=chrome_options)  # you don't have to download chromedriver it will be downloaded by itself and will be saved in cache
    return driver


def RunScrapper(driver):
    start_time = time.time()

    # workbook created
    wb = openpyxl.Workbook()
    # add_sheet is used to create sheet.
    sheet1 = wb.active
    print(" WORKSHEET CREATED SUCCESSFULLY!")
    # INITIALIZING THE COLOUMN NAMES NOW
    c1 = sheet1.cell(row=1, column=1)
    c1.value = "URL"
    c2 = sheet1.cell(row=1, column=2)
    c2.value = "Description"
    c3 = sheet1.cell(row=1, column=3)
    c3.value = "Brand"
    c4 = sheet1.cell(row=1, column=4)
    c4.value = "Product"
    c5 = sheet1.cell(row=1, column=5)
    c5.value = "SKU"
    c6 = sheet1.cell(row=1, column=6)
    c6.value = "Images"
    c7 = sheet1.cell(row=1, column=7)
    c7.value = "Navigation Path"
    c8 = sheet1.cell(row=1, column=8)
    c8.value = "Color"
    c9 = sheet1.cell(row=1, column=9)
    c9.value = "Size"
    c10 = sheet1.cell(row=1, column=10)
    c10.value = "Tags"
    c11 = sheet1.cell(row=1, column=11)
    c11.value = "Regular Price"
    c12 = sheet1.cell(row=1, column=12)
    c12.value = "Sale Price"
    wb.save("smallable.xlsx")
    # setting row number to 2
    mi = 2

    mainlink = "https://www.smallable.com/fr/"
    # driver.get(mainlink)
    # WebDriverWait(driver, 40).until(expected_conditions.visibility_of_element_located((By.XPATH, "//ul[@class='main-nav-ul']")))

    # try:
    #    driver.find_element_by_id('didomi-notice-agree-button').click()
    # except Exception:
    #    pass
    allprods = ['https://www.smallable.com/fr/page/nouveautes', 'https://www.smallable.com/fr/mode/bebe',
                'https://www.smallable.com/fr/mode/enfant', 'https://www.smallable.com/fr/mode/adolescent',
                'https://www.smallable.com/fr/page/streetwear', 'https://www.smallable.com/fr/mode/adulte/femme',
                'https://www.smallable.com/fr/page/beauty',
                'https://www.smallable.com/fr/page/decoration-mobilier-design-enfant',
                'https://www.smallable.com/fr/page/greenable', 'https://www.smallable.com/fr/page/outlet']
    entno = 1
    catno = 1
    for allprod in allprods:
        print("Category No: ", catno)
        catno += 1
        prodlist = []
        try:
            pgno=1
            for j in range(10000000000000000000):
                allprod1 = allprod + "?_page=" + str(pgno)
                driver.get(allprod1)
                print("Page No: ",pgno)
                print("________________________________")
                pgno+=1
                try:
                    WebDriverWait(driver, 6).until(
                    expected_conditions.visibility_of_element_located((By.XPATH, "//div[@class='nbProducts']")))
                    nbpro = driver.find_element_by_xpath("//div[@class='nbProducts']")
                    nopros = nbpro.find_element_by_tag_name('span').text
                    if nopros=='':
                        print("No More Pages")
                        break
                    else:
                        pass
                except Exception:
                    print("No More Pages")
                    break
                if (catno-1) == 1:
                    try:
                        driver.find_element_by_id('didomi-notice-agree-button').click()
                    except Exception:
                        pass
                else:
                    pass
                maincontainer = driver.find_element_by_xpath("//section[@class='product-list']")
                maincontainer = maincontainer.find_elements_by_xpath('.//div[@class="product-item has-ratio"]')
                for maincontain in maincontainer:
                    prodlist.append(maincontain.find_element_by_xpath('.//a[@class="nu product-click"]').get_attribute('href'))

            for prodi in prodlist:
                linktogo = prodi
                print("Entry No: ", entno)
                entno += 1
                print("SCRAPING: ", linktogo)
                try:
                    driver.get(linktogo)

                    WebDriverWait(driver, 40).until(
                        expected_conditions.visibility_of_element_located((By.XPATH, '//div[@class="p-brand"]')))

                    try:
                        a = driver.find_element_by_xpath('//a[@class="hidden-xs"]')
                        tags = a.find_element_by_tag_name('img').get_attribute('title')
                    except Exception:
                        tags = ''
                        pass

                    description = driver.find_element_by_xpath('//div[@class="p-description"]').text
                    navigation = driver.find_element_by_xpath(
                        '//div[@class="hidden-xs c-breadcrumb animate listing-top-bar align-left"]').text

                    colorsites = []
                    colornames = []
                    formcolor = driver.find_element_by_id('form_color_select')
                    colors = formcolor.find_elements_by_tag_name('option')
                    for color in colors:
                        colorsites.append(color.get_attribute('value'))
                        colornames.append(color.text)
                    l = 0
                    for color in colorsites:
                        driver.get(color)
                        WebDriverWait(driver, 40).until(
                            expected_conditions.visibility_of_element_located((By.XPATH, '//div[@class="p-brand"]')))
                        cform = driver.find_element_by_xpath('//div[@class="p-form-color p-form-item"]')
                        try:
                            colorname = cform.find_element_by_xpath('.//div[@class="select-styled"]').text
                        except Exception:
                            colorname = cform.find_element_by_xpath('.//div[@class="noStyle"]').text
                            pass
                        sizeform = driver.find_element_by_xpath('//div[@class="p-form-size p-form-item"]')
                        sizelist = sizeform.find_elements_by_tag_name('li')
                        for sl in range(len(sizelist)):
                            sizeform1 = driver.find_element_by_xpath('//div[@class="p-form-size p-form-item"]')
                            sizeform1.find_element_by_xpath('.//div[@class="select"]').click()
                            sizelist1 = sizeform1.find_elements_by_tag_name('li')
                            try:
                                sizelist1[sl].click()
                            except Exception:
                                pass
                            try:
                                sizename = sizeform1.find_element_by_xpath('.//div[@class="select-styled"]').text
                            except Exception:
                                sizename = sizeform1.find_element_by_xpath('.//div[@class="noStyle active"]').text
                                pass

                            brandm = driver.find_element_by_xpath('//div[@class="p-brand"]')
                            brand = brandm.find_element_by_xpath('.//a[@class="nu"]').text
                            product = driver.find_element_by_xpath('.//div[@class="p-name"]').text

                            try:
                                pricep = driver.find_element_by_xpath('//div[@class="p-price"]')
                                price = pricep.text
                                regularprice = price
                                saleprice = price
                            except Exception:
                                price = ''
                                regularprice = ''
                                saleprice = ''
                                pass

                            try:
                                sku = pricep.find_element_by_tag_name('meta').get_attribute('content')
                            except Exception:
                                sku = ''
                                pass

                            try:
                                imgs = driver.find_elements_by_xpath('//div[@class="image-item"]')
                                imgs1 = driver.find_elements_by_xpath('//div[@class="image-item active"]')
                                imgs = imgs + imgs1
                                x = 1
                                images = ''
                                for img in imgs:
                                    if x == 1:
                                        images = "https:" + img.get_attribute('data-zoom-url')
                                    else:
                                        images = images + ', ' + "https:" + img.get_attribute('data-zoom-url')
                                    x += 1
                            except Exception:
                                images = ''
                                pass

                            print("Brand: ", brand)
                            print("Product: ", product)
                            print("Price: ", price)
                            print("SKU: ", sku)
                            print("Description: ", description)
                            print("Images: ", images)
                            print("Navigation: ", navigation)
                            print("Color Name: ", colorname)
                            print("Size Name: ", sizename)
                            print("Tags: ", tags)
                            c1 = sheet1.cell(row=mi, column=1)
                            c1.value = linktogo
                            c2 = sheet1.cell(row=mi, column=2)
                            c2.value = description
                            c3 = sheet1.cell(row=mi, column=3)
                            c3.value = brand
                            c4 = sheet1.cell(row=mi, column=4)
                            c4.value = product
                            c5 = sheet1.cell(row=mi, column=5)
                            c5.value = sku
                            c6 = sheet1.cell(row=mi, column=6)
                            c6.value = images
                            c7 = sheet1.cell(row=mi, column=7)
                            c7.value = navigation
                            c8 = sheet1.cell(row=mi, column=8)
                            c8.value = colorname
                            c9 = sheet1.cell(row=mi, column=9)
                            c9.value = sizename
                            c10 = sheet1.cell(row=mi, column=10)
                            c10.value = tags
                            c11 = sheet1.cell(row=mi, column=11)
                            c11.value = regularprice
                            c12 = sheet1.cell(row=mi, column=12)
                            c12.value = saleprice
                            mi += 1
                            print("-----------------------------------")
                    wb.save("smallable.xlsx")
                except Exception:
                    print("Broken Link")
                print("")
                print("**************************************************************")
                print("")
        except Exception:
            print("Broken Category")
            pass
        print("")
        print("**************************************************************")
        print("")
        print("")
        print("**************************************************************")
        print("")

    # give time taken to execute everything
    print("time elapsed: {:.2f}s".format(time.time() - start_time))


# create the driver object.
driver = configure_driver()

# call the scrapper to run
RunScrapper(driver)

# close the driver.
# driver.close()














